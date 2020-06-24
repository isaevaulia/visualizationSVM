import openpyxl as xl
import copy
import numpy as np
from UtilityClasses import *
from matplotlib import markers
from matplotlib.lines import Line2D
import Constants
from PyQt5.QtWidgets import QTableWidgetItem
import matplotlib.pyplot as plt

""" Некоторые вспомогательные функции вынесены сюда.
"""

def readExcelData(filename):
    """ Читает excel файл.

    :param filename: имя файла
    :return: возвращает массив объектов типа Column. Колонки прочтенной таблицы
    """
    columns = []
    workbook = xl.load_workbook(filename=filename)
    worksheet = workbook.get_active_sheet()
    columnData = []
    target = []
    columnName = None
    for i, col in enumerate(worksheet.iter_cols(min_col=1, max_col=worksheet.max_column)):
        for cell in col:
            if cell.row == 1:
                columnName = cell.internal_value
            else:
                columnData.append(cell.internal_value)
        column = Column(columnName, columnData, i)
        if column.getName() == 'target':
            target = column
        else:
            columns.append(column)
        columnData = []
    return columns, target

"""moe"""
def drawPlane(w,w0, canvas, xmin=1,xmax=5,color='r-'):
    """ рисует разделяющую гиперплоскость"""
    figure = canvas.figure
    axes = figure.add_subplot(1, 1, 1)
    def f(x):
        return (w0-w[0]*x)/w[1]
    y= np.vectorize(f)
    x= np.linspace(xmin,xmax,50)
    axes.plot(x,y(x),color)

    canvas.draw()

def make_meshgrid(xPoint, h=.02):
    x = []
    y = []
    for i in enumerate(xPoint):
        x.append(i[1][0])
        y.append(i[1][1])
    x_min, x_max = min(x) - 1, max(x) + 1
    y_min, y_max = min(y) - 1, max(y) + 1
    xx, yy = np.meshgrid(np.arange(x_min, x_max, h), np.arange(y_min, y_max, h))
    return xx, yy

def plot_contours(ax, clf, xx, yy, **params):
    Z = clf.predict(np.c_[xx.ravel(), yy.ravel()])
    Z = Z.reshape(xx.shape)
    out = ax.contourf(xx, yy, Z, **params)
    return out

def drawContours(clf, canvas, x):
    figure = canvas.figure
    axes = figure.add_subplot(1, 1, 1)
    xx, yy = make_meshgrid(x)
    plot_contours(axes, clf, xx, yy, cmap=plt.cm.coolwarm, alpha=0.8)
    canvas.draw()

def drawBrokenLine(points, canvas):
    """ Рисует ломаную линию.

    :param points: массив объектов типа Point. Множество точек, из которых состоит ломаная. Порядок важен.
    :param canvas: канвас, куда происходит рисование
    """
    figure = canvas.figure
    axes = figure.add_subplot(1, 1, 1)
    xpolygonData = []
    ypolygonData = []
    for point in points:
        xpolygonData.append(point.getX())
        ypolygonData.append(point.getY())
    if len(points) == 1:
        axes.plot(xpolygonData, ypolygonData, linestyle="None", marker="o", color='b', markersize=4)
    else:
        axes.plot(xpolygonData, ypolygonData, "b")
    canvas.draw()

def drawPolygon(polygon, canvas):
    """ Рисует многоугольник.

    :param polygon: массив объектов типа Point. Множество точек из которых состоит многоугольник. Порядок важен.
    :param canvas: канвас, куда происходит рисование
    """
    newPolygonContainter = copy.deepcopy(polygon)
    lastPoint = Point(polygon[0].getX(), polygon[0].getY(), None)
    newPolygonContainter.append(lastPoint)
    drawBrokenLine(newPolygonContainter, canvas)

def crossingNumberAlgorithm(point, polygon):
    """ Здесь реализован алгоритм, определяющий, лежит ли точка внутри многоугольника.

    :param point: объект типа Point. Проверяемая точка.
    :param polygon: массив объектов типа Point. Проверяемый многоугольник
    :return: возвращает True если точка лежит внутри многоугольника. В противном случае возвращает False.
    """
    intersectionCounter = 0
    if len(polygon) < 3:
        return False
    A1 = 0
    B1 = 1
    C1 = - point.getY()
    for i in range(0, len(polygon)):
        if i == len(polygon) - 1:
            X1 = polygon[i].getX()
            Y1 = polygon[i].getY()
            X2 = polygon[0].getX()
            Y2 = polygon[0].getY()
        else:
            X1 = polygon[i].getX()
            Y1 = polygon[i].getY()
            X2 = polygon[i + 1].getX()
            Y2 = polygon[i + 1].getY()
        A2 = Y2 - Y1
        B2 = - (X2 - X1)
        C2 = (X2 - X1) * Y1 - (Y2 - Y1) * X1
        Y = point.getY()
        if A2 != 0:
            X = (- C2 - B2*Y) / A2
            if Y1 > Y2:
                var = Y2
                Y2 = Y1
                Y1 = var
            if (X > point.getX()) and (Y < Y2) and (Y > Y1):
                intersectionCounter += 1
    if intersectionCounter % 2 == 0:
        return False
    else:
        return True

def getFlippedMarkerDictionary():
    """ Возня с маркерами в matplotlib. Ничего особенного, хоть и ничего не понятно.

    """
    newDict = {}
    for key, value in markers.MarkerStyle.markers.items():
        newDict[value] = key
    return newDict

def getSupportiveLine(columns, i, j):
    """ Настоящее магическое место в этой программе.
    Объяснить что тут происходит сложно. После долгих ковыряний в matplotlib, выяснилось, что
    в matplotlib график состоит из многих линий line. line может представлять из себя несколько отрезков или
    несколько точек. При каждом рисовании на графике создается новый line. Для того чтобы выделение точек работало -
    нужно знать координаты всех точек, изображенных на этом графике. Короче. Прежде чем отрисовывать кластеры -
    я рисую все точки какие есть и делаю их невидимыми. После, поверх всего этого рисуем сначала точки,
    не принадлежащие никакому кластеру. См. функцию getDummyCluster. А потом рисуем все кластеры.
    Теперь координаты всех точек, изображенных на графике можно достать из нулевой линии. И делается это так:
    self._xData = axes.lines[0].get_xdata()
    self._yData = axes.lines[0].get_ydata()

    :param columns: Массив объектов типа Column. Исходные данные.
    :param i: Ось X в проекции, для которой вызвана эта функция
    :param j: Ось Y в проекции, для которой вызвана эта функция
    :return: Объект типа Line2D из библиотеки matplotlib. Все точки белым цветом.
    """

    line = Line2D(columns[i], columns[j],
                  linestyle="None",
                  marker=Constants.DEFAULT_POINT_SHAPE,
                  color=Constants.INVISIBLE_COLOR,
                  markersize=Constants.DEFAULT_MARKER_SIZE_SMALL)
    return line

def fillTableWithData(tablewidget, data):
    """ Функция заполняет таблицу данными.

    :param tablewidget: PyQt виджет - заполняемая таблица
    :param data: Объект типа Data - данные для заполнения.
    """
    tablewidget.setColumnCount(data.columnCount())
    tablewidget.setRowCount(data.rowCount())
    horizontalheaders = data.getColumnNames()
    verticalheaders = []
    for i, row in enumerate(data):
        verticalheaders.append(str(row.getIndex()))
        for j, element in enumerate(row):
            tablewidget.setItem(i, j, QTableWidgetItem(str(element)))
    tablewidget.setHorizontalHeaderLabels(horizontalheaders)
    tablewidget.setVerticalHeaderLabels(verticalheaders)

def fillTableWithCluster(tablewidget, cluster):
    """ Заполняет таблицу данными из кластера.

    :param tablewidget: PyQt виджет - заполняемая таблица
    :param cluster: Объект типа Cluster - данные для заполнения.
    """
    if cluster.getSize() != 0:
        fillTableWithData(tablewidget, cluster.getInnerData())





